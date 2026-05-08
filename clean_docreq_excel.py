"""
clean_docreq_excel.py
Cleans the DNV Document Tracker Excel file for import into the DocReq system.

DNV DocReq sheet:
  - Rows 1-5 (title block + header row): untouched
  - Rows 6+: columns A-F kept, columns G onwards cleared

All other sheets:
  - First non-empty row (header): kept
  - All data rows below: cleared

Writes a clean copy, then re-parses it into data/dnv_doc_requirements.csv.
Also deletes state/docreq.db so the app re-seeds clean on next load.
"""
import csv
import shutil
from pathlib import Path

import openpyxl
import pandas as pd

SRC   = Path("C:/Users/lchas/Downloads/DNV_Document_Tracker - 2.0 (9) (1).xlsx")
CLEAN = Path("C:/Users/lchas/Downloads/DNV_Document_Tracker_CLEAN.xlsx")
CSV_OUT = Path("C:/Users/lchas/DNV _System/data/dnv_doc_requirements.csv")
DB_PATH = Path("C:/Users/lchas/DNV _System/state/docreq.db")

DOCREQ_SHEET  = "DNV DocReq"
DOCREQ_HEADER_ROW = 5   # 1-indexed row with column headers
KEEP_COLS     = 6        # keep columns A-F (1-6 in openpyxl)


# ── Step 1: Clean the Excel ───────────────────────────────────────────────────
def clean_excel():
    print(f"Loading {SRC.name} ...")
    wb = openpyxl.load_workbook(str(SRC))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name == DOCREQ_SHEET:
            # Keep rows 1-5 intact (title block + header)
            # For rows 6+: clear columns G onwards (col 7+)
            cleared = 0
            for row in ws.iter_rows(min_row=DOCREQ_HEADER_ROW + 1,
                                     max_row=ws.max_row,
                                     min_col=KEEP_COLS + 1,
                                     max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.value = None
                        cleared += 1
            print(f"  {sheet_name}: cleared {cleared} cells in columns G+ (kept A-F + header)")

        else:
            # Find the first non-empty row (header)
            header_row = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                if any(c is not None for c in row):
                    header_row = i
                    break

            if header_row is None:
                print(f"  {sheet_name}: no header found, skipping")
                continue

            # Clear all rows below the header
            cleared = 0
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        cell.value = None
                        cleared += 1
            print(f"  {sheet_name}: header kept at row {header_row}, cleared {cleared} data cells")

    wb.save(str(CLEAN))
    print(f"\nClean file saved: {CLEAN.name}")


# ── Step 2: Re-parse DocReq sheet into CSV ────────────────────────────────────
def _clean_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\n", " ").replace("\r", " ")
    if s.lower() in ("nan", "none", "0"):
        return ""
    return s


def reparse_csv():
    print(f"\nParsing {CLEAN.name} -> dnv_doc_requirements.csv ...")
    df = pd.read_excel(str(CLEAN), sheet_name=DOCREQ_SHEET, header=DOCREQ_HEADER_ROW - 1)

    # Col A (index 0) is blank/unnamed — skip it
    # Col B (index 1) = Rules Table Ref.
    # Col C (index 2) = Item number
    # Col D (index 3) = Code
    # Col E (index 4) = Object/description
    # Col F (index 5) = Additional description
    col_map = {
        df.columns[1]: "rule_ref",
        df.columns[2]: "item_number",
        df.columns[3]: "dnv_code",
        df.columns[4]: "requirement_text",
        df.columns[5]: "additional_description",
    }

    df = df.rename(columns=col_map)

    rows = []
    current_item = None

    for _, r in df.iterrows():
        item = _clean_str(r.get("item_number", ""))
        code = _clean_str(r.get("dnv_code", ""))
        text = _clean_str(r.get("requirement_text", ""))
        desc = _clean_str(r.get("additional_description", ""))
        rule = _clean_str(r.get("rule_ref", ""))

        # Skip legend rows (code starts with '=' or 'Code*') and blanks
        if code.startswith("=") or code == "Code*":
            continue
        if not item and not code and not text:
            continue

        # Numeric item numbers come as floats — normalise
        if item:
            try:
                item = str(int(float(item)))
            except (ValueError, OverflowError):
                pass
            current_item = item

        rows.append({
            "item_number":           item,
            "rule_ref":              rule or "Pt.5 Ch.10 Sec.1",
            "dnv_code":              code,
            "requirement_text":      text,
            "additional_description": desc,
            "approval_type":         "",   # cleared — will be entered via UI
            "document_number":       "",
            "document_description":  "",
            "owner":                 "",
            "dnv_status":            "",
            "open_comments":         0,
            "subsystem":             "",
            "coverage_status":       "MISSING",
        })

    df_out = pd.DataFrame(rows)
    df_out.to_csv(str(CSV_OUT), index=False)

    numbered = df_out[df_out["item_number"] != ""]
    print(f"  Written {len(df_out)} rows ({len(numbered)} numbered requirements)")
    return len(numbered)


# ── Step 3: Delete old DB so app re-seeds ────────────────────────────────────
def reset_db():
    if DB_PATH.exists():
        DB_PATH.unlink()
        print(f"\nDeleted {DB_PATH.name} — app will re-seed on next /docreq load")
    else:
        print(f"\n{DB_PATH.name} not found — nothing to delete")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    clean_excel()
    count = reparse_csv()
    reset_db()
    print(f"\nDone. {count} clean requirements ready.")
    print("Restart Flask (or just navigate to /docreq) to re-seed the database.")
