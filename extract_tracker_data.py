"""
DNV_Document_Tracker 2.0 — Extraction Script
Produces three new tables for the DNV Coordinator app:
  1. dnv_comments.csv      — 1,271 deduplicated DNV comments/TQs
  2. document_register.csv — 482 documents across 9 subsystem sheets
  3. Appends 72 TBC action items to action_tracker.csv
Also updates config.json to register the new tables.
"""
import sys, os, csv, json, shutil
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

SRC     = r"C:\Users\lchas\Downloads\DNV_Document_Tracker - 2.0 (9).xlsx"
DATA    = r"C:\Users\lchas\DNV _System\data"
CONFIG  = r"C:\Users\lchas\DNV _System\config.json"
BACKUP  = os.path.join(r"C:\Users\lchas\DNV _System\_backups",
                       f"backup_{datetime.now().strftime('%Y-%m-%d_%H%M')}_tracker_extract")

os.makedirs(BACKUP, exist_ok=True)
for f in ["action_tracker.csv", "config.json"]:
    src = os.path.join(DATA, f) if f != "config.json" else CONFIG
    if os.path.exists(src):
        shutil.copy2(src, os.path.join(BACKUP, f))
print(f"Backup: {BACKUP}\n")

import openpyxl
wb = openpyxl.load_workbook(SRC, data_only=True)

def fmt_date(v):
    if v is None: return ""
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    return str(v).strip()

def clean(v):
    if v is None: return ""
    s = str(v).strip().replace("\n", " ").replace("\r", "")
    return s

# ═══════════════════════════════════════════════════════════════════════════════
# 1. DNVComments → dnv_comments.csv
#    Dedup on CommentId (keep first occurrence = unique comment record)
#    Aggregate all DocumentNo / DocumentTitle per CommentId as pipe list
#    Keep 13 signal columns; drop all 23 computed Excel formula columns
# ═══════════════════════════════════════════════════════════════════════════════
print("Extracting DNVComments...")

ws = wb["DNVComments"]
# Column indices (0-based) for the signal columns we want
# Col 1=Status, 2=CommentId, 5=Type, 6=Title, 7=Description,
# 10=Discipline, 11=RuleReg, 12=IssueDate, 14=IssuedByUserName,
# 17=ClosedDate, 20=ActionNeededBy, 23=DocumentNo, 25=DocumentTitle,
# 47=DEEPInternalNotes, 48=ResponsibleParty, 51=ActionNeeded
SIG = {
    "Status":            1,
    "CommentID":         2,
    "Type":              5,
    "Title":             6,
    "Description":       7,
    "Discipline":        10,
    "RuleRef":           11,
    "IssueDate":         12,
    "IssuedBy":          14,
    "ClosedDate":        17,
    "ActionNeededBy":    20,
    "DEEPInternalNotes": 47,
    "ResponsibleParty":  48,
    "ActionNeeded":      51,
}
# Document aggregation columns
DOC_NO_COL    = 23
DOC_TITLE_COL = 25

# First pass: collect all doc numbers/titles per CommentId
doc_map = {}   # CommentId → set of (docno, doctitle)
first_row = {} # CommentId → first data row dict

for row in ws.iter_rows(min_row=2, values_only=True):
    cid = row[2]
    if cid is None:
        continue
    doc_no    = clean(row[DOC_NO_COL])
    doc_title = clean(row[DOC_TITLE_COL])
    if cid not in doc_map:
        doc_map[cid]   = []
        first_row[cid] = row
    if doc_no and doc_no not in [d[0] for d in doc_map[cid]]:
        doc_map[cid].append((doc_no, doc_title))

# Build output rows
DNV_FIELDS = list(SIG.keys()) + ["LinkedDocuments", "LinkedDocTitles"]
dnv_rows = []

for cid, row in first_row.items():
    rec = {}
    for col_name, col_idx in SIG.items():
        v = row[col_idx]
        if col_name in ("IssueDate", "ClosedDate", "ActionNeededBy"):
            rec[col_name] = fmt_date(v)
        else:
            rec[col_name] = clean(v)
    docs   = doc_map.get(cid, [])
    rec["LinkedDocuments"] = " | ".join(d[0] for d in docs if d[0])
    rec["LinkedDocTitles"] = " | ".join(d[1] for d in docs if d[1])
    dnv_rows.append(rec)

# Sort by IssueDate desc so newest comments appear first
dnv_rows.sort(key=lambda r: r["IssueDate"] or "0000", reverse=True)

out_path = os.path.join(DATA, "dnv_comments.csv")
with open(out_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=DNV_FIELDS)
    w.writeheader()
    w.writerows(dnv_rows)

open_c   = sum(1 for r in dnv_rows if r["Status"] == "Open")
closed_c = sum(1 for r in dnv_rows if r["Status"] == "Closed")
other_c  = len(dnv_rows) - open_c - closed_c
print(f"  Written: dnv_comments.csv — {len(dnv_rows)} unique comments")
print(f"  Open: {open_c}  |  Closed: {closed_c}  |  Other: {other_c}")

# Discipline breakdown of open comments
from collections import Counter
open_by_disc = Counter(r["Discipline"] for r in dnv_rows if r["Status"] == "Open")
print("  Open comments by discipline:")
for disc, cnt in open_by_disc.most_common():
    print(f"    {disc or '(unset)'}: {cnt}")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. Subsystem sheets → document_register.csv
#    Stack 9 sheets, add Subsystem column, strip empty/header rows
# ═══════════════════════════════════════════════════════════════════════════════
print("\nConsolidating subsystem document sheets...")

SUBSYSTEM_SHEETS = [
    "Parent", "Viewports", "Baseplate", "Tanks",
    "Power", "Comms", "GasAux", "Structures", "HSB",
]
# Common schema across all sheets (first 16 cols)
REG_FIELDS = [
    "Subsystem", "Description", "VendorDocNumber", "Owner", "Version",
    "SubmittedDate", "LastActionDate", "Compliance", "Comments", "Actions",
    "DDMNumber", "VeracityCategory", "DocumentComplete", "DNVSubmitted",
    "DNVStatus", "OpenComments", "OpenUniqueComments",
]

reg_rows = []
for sheet_name in SUBSYSTEM_SHEETS:
    ws2 = wb[sheet_name]
    sheet_rows = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        # Skip rows where Description (col 0) is None/empty
        if not row[0]:
            continue
        rec = {
            "Subsystem":         sheet_name,
            "Description":       clean(row[0]),
            "VendorDocNumber":   clean(row[1]),
            "Owner":             clean(row[2]),
            "Version":           clean(row[3]),
            "SubmittedDate":     fmt_date(row[4]),
            "LastActionDate":    fmt_date(row[5]),
            "Compliance":        clean(row[6]),
            "Comments":          clean(row[7]),
            "Actions":           clean(row[8]),
            "DDMNumber":         clean(row[9]),
            "VeracityCategory":  clean(row[10]),
            "DocumentComplete":  clean(row[11]),
            "DNVSubmitted":      clean(row[12]),
            "DNVStatus":         clean(row[13]),
            "OpenComments":      clean(row[14]),
            "OpenUniqueComments": clean(row[15]),
        }
        reg_rows.append(rec)
        sheet_rows += 1
    print(f"  {sheet_name:<15}: {sheet_rows} documents")

out_path2 = os.path.join(DATA, "document_register.csv")
with open(out_path2, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=REG_FIELDS)
    w.writeheader()
    w.writerows(reg_rows)

total_open_docs = sum(1 for r in reg_rows if r["OpenComments"] and r["OpenComments"] != "0")
total_unapproved = sum(1 for r in reg_rows if r["DNVStatus"] not in ("Approved", "Appr w/ Comm", "For Info", "Discarded", "") and r["DNVStatus"])
print(f"  Written: document_register.csv — {len(reg_rows)} total documents")
print(f"  Documents with open DNV comments: {total_open_docs}")

# ═══════════════════════════════════════════════════════════════════════════════
# 3. TBCs → append to action_tracker.csv
#    Left column: missing documents → ACT-VAN-M-001...
#    Right column: revision needed  → ACT-VAN-R-001...
# ═══════════════════════════════════════════════════════════════════════════════
print("\nExtracting TBCs...")

ws3 = wb["TBCs"]
tbc_rows = list(ws3.iter_rows(min_row=2, values_only=True))  # skip header

ACT_FIELDS = ["ActionID","Title","Owner","DueDate","Status","RAGStatus",
              "LastUpdated","Category","Notes"]

# Read existing action tracker
act_path = os.path.join(DATA, "action_tracker.csv")
with open(act_path, encoding="utf-8") as f:
    existing_actions = list(csv.DictReader(f))

new_actions = []
today = "2026-04-24"

# Left column: Missing documents (cols 0-5: idx, description, ddm, status, owner, notes)
missing_count = 0
for row in tbc_rows:
    idx   = row[0]
    desc  = clean(row[1])
    ddm   = clean(row[2])
    status_note = clean(row[3])
    owner = clean(row[4])
    notes = clean(row[5])
    if not idx or not desc:
        continue
    missing_count += 1
    act_id = f"ACT-VAN-M-{missing_count:03d}"
    # Infer RAG from status_note
    rag = "AMBER"
    if status_note.lower() in ("in work", "in approval"):
        rag = "AMBER"
    elif not status_note:
        rag = "RED"
    note_parts = []
    if ddm:        note_parts.append(f"DDM: {ddm}")
    if status_note: note_parts.append(f"Status: {status_note}")
    if notes:      note_parts.append(notes)
    note_str = "  |  ".join(note_parts) if note_parts else "No DDM assigned. Missing from DNV submission package."
    new_actions.append({
        "ActionID":    act_id,
        "Title":       f"[MISSING DOC] {desc}",
        "Owner":       owner or "TBD",
        "DueDate":     "",
        "Status":      "Open",
        "RAGStatus":   rag,
        "LastUpdated": today,
        "Category":    "Technical",
        "Notes":       f"Source: DNV_Document_Tracker TBCs — Missing Documents list. {note_str}",
    })

print(f"  Missing documents: {missing_count} action items")

# Right column: Documents needing revision (cols 7-11: idx, description, ddm, owner, notes)
revision_count = 0
for row in tbc_rows:
    idx   = row[7]
    desc  = clean(row[8])
    ddm   = clean(row[9])
    owner = clean(row[10])
    notes = clean(row[11])
    if not idx or not desc:
        continue
    revision_count += 1
    act_id = f"ACT-VAN-R-{revision_count:03d}"
    note_parts = []
    if ddm:   note_parts.append(f"DDM: {ddm}")
    if notes: note_parts.append(notes)
    note_str = "  |  ".join(note_parts) if note_parts else "Revision required before DNV resubmission."
    new_actions.append({
        "ActionID":    act_id,
        "Title":       f"[REVISION REQ] {desc}",
        "Owner":       owner or "TBD",
        "DueDate":     "",
        "Status":      "Open",
        "RAGStatus":   "AMBER",
        "LastUpdated": today,
        "Category":    "Technical",
        "Notes":       f"Source: DNV_Document_Tracker TBCs — Documents Needing Revision. {note_str}",
    })

print(f"  Revision required: {revision_count} action items")

# Write combined action tracker
all_actions = existing_actions + new_actions
with open(act_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=ACT_FIELDS)
    w.writeheader()
    w.writerows(all_actions)

print(f"  action_tracker.csv updated: {len(existing_actions)} existing + {len(new_actions)} new = {len(all_actions)} total")

# ═══════════════════════════════════════════════════════════════════════════════
# 4. Update config.json — register new tables
# ═══════════════════════════════════════════════════════════════════════════════
print("\nUpdating config.json...")

with open(CONFIG, encoding="utf-8") as f:
    cfg = json.load(f)

cfg["csv_files"]["dnv_comments"]       = "dnv_comments.csv"
cfg["csv_files"]["document_register"]  = "document_register.csv"

# Update critical path — add Baseplate (highest doc count, most open comments)
cfg["critical_path_docs"] = [
    "BT3345_9001_009", "SDD-SENTINEL-001", "FMECA-SENTINEL-001",
    "VA2-0003300",  # Baseplate Design Basis
    "VA2-0003299",  # Inspection & Test Plan (on TBCs revision list)
    "VA2-0003416",  # FMECA (on TBCs missing list)
]

with open(CONFIG, "w", encoding="utf-8") as f:
    json.dump(cfg, f, indent=2)
print("  config.json updated — dnv_comments and document_register registered")

# ═══════════════════════════════════════════════════════════════════════════════
# Summary
# ═══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*62)
print("EXTRACTION COMPLETE")
print("="*62)
print(f"  dnv_comments.csv      : {len(dnv_rows):>4} unique comments "
      f"({open_c} open / {closed_c} closed / {other_c} other)")
print(f"  document_register.csv : {len(reg_rows):>4} documents across "
      f"{len(SUBSYSTEM_SHEETS)} subsystems")
print(f"  action_tracker.csv    : {len(all_actions):>4} total actions "
      f"({missing_count} missing-doc + {revision_count} revision-req added)")
print(f"\n  Backup : {BACKUP}")

# Quick discipline / status breakdown for the report
print("\n  DNV Comment open breakdown by discipline:")
for disc, cnt in open_by_disc.most_common(10):
    bar = "=" * min(cnt // 2, 30)
    label = str(disc) if disc else "(unset)"
    print(f"    {label:<40} {cnt:>3}  {bar}")
